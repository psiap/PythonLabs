#2.	Напишите скрипт для информационной системы библиотеки. 
#База данных библиотеки включает таблицы «Авторы» с полями «id», 
#«имя», «страна», «годы жизни», и «Книги» с полями «id автора», 
#«название», «количество страниц», «издательство», «год издания»). 
#Необходимо производить авторизацию пользователей, логины и пароли
# которых хранятся в отдельной таблице. Пароли должны храниться в 
# зашифрованном виде (например, хэш SHA-1 или MD5). В программе должны 
# быть окна для отображения информации о всех книгах и авторах, окно 
# добавления книги/автора. Реализуйте также возможность сохранения 
# информации о выделенном авторе в файле в формате json или XML (по выбору пользователя). 
# При добавлении нового автора в базу допускается не заполнять поля в соответствующем окне, 
# а распарсить файл, указанный пользователем (файл необходимо заранее создать и заполнить 
# информацией вручную, в текстовом редакторе). Для преобразования в формат XML и json напишите 
# собственный код; парсинг можно делать с помощью сторонних библиотек. Форматы файлов: 
 
#JSON: 
#{ 
#  "name": "L.N.Tolstoi", 
#  "country": "Russia", 
#  "years": [1828, 1910] 
#} 	XML: 
 
#<author> 
#  <name>L.N.Tolstoi</Name> 
#  <country>Russia</Country> 
#  <years born=”1828” died=”1910”/> 
#</author>  


import sys
import openpyxl
import xlrd, xlwt
import hashlib
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *

class Example(QWidget):

    def __init__(self):
        super().__init__()
         # Создадим виджет и сохраним ссылку на него

        
        

        self.initUI()

        #Основное окно
    def initUI(self):
        #Первые два параметра х и у - это позиция окна. Третий - ширина, и четвертый - высота окна
        self.setGeometry(300, 300, 250, 100)
        self.setWindowTitle('Library_login')
        self.setWindowIcon(QIcon('web.jpg'))

        self.label1 = QLabel("Логин", self)
        self.label1.setGeometry(13,0,110,20)
        
        self.label2 = QLabel("Пароль", self)
        self.label2.setGeometry(13,30,110,20)

        self.btn = QPushButton('Войти', self)
        self.btn.setGeometry(50,60,110,20)

        self.text_login = QLineEdit(self)
        self.text_login.setGeometry(50,0,110,20)

        self.text_password = QLineEdit(self)
        self.text_password.setGeometry(50,30,110,20)

        self.btn.clicked.connect(self.on_click)
        
        self.show()

    @pyqtSlot()
    def on_click(self):
        la = authorization(self.text_login.text(),self.text_password.text())
        print("Get::", la.exel_get()," Login::",self.text_login.text()  ," Password::",self.text_password.text())
        self.text_login.setText("")
        self.text_password.setText("")

        if la.exel_get() == 1:
            self.dialog = Dialog()
        
        

class Dialog(QWidget):
    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):
        self.listWidget = QListWidget()
        self.listWidget.setGeometry(300, 300, 550, 400)
        self.listWidget.setWindowTitle('Library')
        self.listWidget.setWindowIcon(QIcon('web.jpg'))
        g = parse_librery("авторы")
        
        
        self.listWidget.addItem(g.pars_list())
        self.btn = QPushButton('Войти', self)
        self.btn.setGeometry(50,60,110,20)
        
        self.listWidget.show()
        self.btn.show()

            #Парсер таблицы
class parse_librery(object):
    def __init__(self, list_s):
        self.__list_s = list_s

    def pars_list(self):
        self.wb = openpyxl.load_workbook('library.xlsx')
        self.sheet = self.wb[self.__list_s]
        self.wb.active = 0
        self.rows = self.sheet.max_row
        self.cols = self.sheet.max_column
        for i in range(1, self.rows + 1):
              self.string = ''
              for j in range(1, self.cols + 1):
                  self.cell = self.sheet.cell(row = i, column = j)
                  self.string = self.string + str(self.cell.value) + ' '
              print(self.string)

        return self.string

    def set_json(self):
        self.name = 'Володя'
        self.country = 'Абра'
        self.years = '2999'
        return "JSON:\n" +'{\n  "name": ' + '"' + self.name + '",\n' + '  "country": ' + '"' + self.country + '",\n' + '  "years": ' + "[" + self.years + "]\n}"  

    def set_xml(self):
        self.name = 'Володя'
        self.country = 'Абра'
        self.years = '2999'
        return "XML:\n\n" + "<author>\n  <name>" + self.name + "</name>\n  <country>" + self.country + '</country>\n  <years born="' + self.years + '" died="' + self.years + '"/>\n</author>'

        


        #Login
class authorization(object):

    def __init__(self, login, password):
        self.__login = login
        self.__password = password
        self.__login = authorization.computeMD5hash(self.__login)
        self.__password = authorization.computeMD5hash(self.__password)

    def exel_get(self):
        try:
         self.__rb = xlrd.open_workbook('password.xlsx')
         self.__sheet = self.__rb.sheet_by_index(0)
         self.__count = 0 
         while(self.__sheet.row_values(self.__count)[0] != 'None'):

            if authorization.computeMD5hash(self.__sheet.row_values(self.__count)[0]) == self.__login:
                 self.__count == 0
                 while(self.__sheet.row_values(self.__count)[1] != 'None'):
                     
                    if authorization.computeMD5hash(self.__sheet.row_values(self.__count)[1]) == self.__password:
                        return 1
                    self.__count += 1
            elif self.__sheet.row_values(self.__count)[0] == 'None':
                return 0

               
            self.__count += 1
        except Exception:
            return 0
        
    @staticmethod
    def computeMD5hash(my_string):
        m = hashlib.md5()
        m.update(my_string.encode('utf-8'))
        return m.hexdigest()

if __name__ == '__main__':
    
    app = QApplication(sys.argv)
    ex = Example()
    sys.exit(app.exec_())
 




